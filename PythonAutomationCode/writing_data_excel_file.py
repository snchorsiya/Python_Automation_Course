import openpyxl

# file > workbook > sheet > cells

file_path = "C:\\Users\\Sheetal.Chorsiya\\Downloads\\MyBlackSheet.xlsx"
workbook = openpyxl.load_workbook(file_path)
# sheet = workbook["Sheet1"]
sheet = workbook.active

# for r in range(1, 8):
#     for c in range(1, 4):
#         sheet.cell(r, c).value = "Hello"

sheet.cell(1, 1).value = "Name"
sheet.cell(1, 2).value = "Age"
sheet.cell(1, 3).value = "ID"

sheet.cell(2, 1).value = "Sheetal"
sheet.cell(2, 2).value = "26"
sheet.cell(2, 3).value = "1"

sheet.cell(3, 1).value = "Raj"
sheet.cell(3, 2).value = "16"
sheet.cell(3, 3).value = "2"


workbook.save(file_path)
