import openpyxl

# file > workbook > sheet > cells
file_path = "C:\\Users\\Sheetal.Chorsiya\\Downloads\\MyTestFile.xlsx"
workbook = openpyxl.load_workbook(file_path)
# sheet = workbook.active  # only access single sheet
sheet = workbook["MyEmpsList"]

rows = sheet.max_row
cols = sheet.max_column
print("Number of rows containing data are:", rows)
print("Number of columns containing data are:", cols)

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r, c).value, end=" ")
    print(" ")

sheet1 = workbook["MyFruitList"]
row = sheet1.max_row
col = sheet1.max_column
print(row)
print(col)

for ro in range(1, row+1):
    for co in range(1, col+1):
        print(sheet1.cell(ro, co).value, end=" ")
    print(" ")